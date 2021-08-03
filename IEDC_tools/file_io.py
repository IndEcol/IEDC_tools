"""
Functions for file input-output operations.
"""

import os

import openpyxl
import numpy as np
import pandas as pd

import IEDC_paths
from IEDC_tools import dbio


def read_input_file(file):
    """
    Function to load the master classification file and return its content as a DataFrame.

    :return: Master classification data as pandas dataframe
    """
    df = pd.read_excel(file.file, sheet_name=file.sheet, header=file.header, index_col=file.index)
    df = df.drop(file.drop_rows, axis=0)
    df = df.drop(file.drop_rows, axis=0)
    return df


def get_candidate_filenames(path=IEDC_paths.candidates, verbose=False):
    """
    Browses a directory and returns the filenames as a list.

    :param path: The directory of the files to be scanned
    :param verbose: Print some info. Options are 1 (count) and 2 (all names)
    :return: List of filenames
    """
    # Let's exclude temporary and hidden files
    # TODO: Will probably need more love for Windows...
    exclude_first_letter = ['.', '~']
    files = os.listdir(path)
    # discard hidden files, etc.
    files = [f for f in files if not any([f.startswith(ex) for ex in exclude_first_letter])]
    # only consider Excel files
    files = [f for f in files if f.endswith('.xlsx')]
    if verbose == 1:
        print("Found %s candidate files in %s" % (len(files), path))
    elif verbose == 2:
        print("Found the following %s candidate files: %s" % (len(files), files))
    return files


def read_candidate_meta(file, path=IEDC_paths.candidates):
    """
    Will read a candidate file and return its metadata.

    The hardcoding of `skiprows` and `usecols` is ugly, but necessary with the current data template. Something to
    consider for the next template.

    :param file: Filename of the file to process
    :param path: Path of the file
    :return: Dictionary of dataframes for metadata, row_classifications, and data
    """
    # make it a proper path
    file = os.path.join(path, file)
    # Check what type of file this is, i.e. LIST or TABLE formatted data
    workbook = openpyxl.load_workbook(file, data_only=True)
    worksheet = workbook['Cover']
    data_type = worksheet.cell(10, 7).value  # i.e. cell G10
    # Get dataset information table on Cover sheet
    dataset_info = pd.read_excel(file, sheet_name='Cover', usecols='C:D',
                         skiprows=[0, 1], index_col="Column name")
    # The `data_sources` metadata should be in the same position for both TABLE and LIST, i.e. F6:H9
    data_sources = pd.read_excel(file, sheet_name='Cover', usecols='F:H',
                                 skiprows=[i for i in range(4)], index_col=0, nrows=5,
                                 header=None, names=['i', 'a', 'b'])
    # Excel templates should be unified for both types in the future :(
    if data_type == 'TABLE':
        row_classifications = pd.read_excel(file, sheet_name='Cover', usecols='F:G',
                                            skiprows=[i for i in range(10)],
                                            index_col="Row Aspects classification").dropna()
        col_classifications = pd.read_excel(file, sheet_name='Cover', usecols='H:I',
                                            skiprows=[i for i in range(10)],
                                            index_col="Col Aspects classification").dropna()
        data_info = pd.read_excel(file, sheet_name='Cover', usecols='J:K',
                                  skiprows=[i for i in range(10)],
                                  index_col="DATA").dropna()
        u_nominator = worksheet.cell(7, 8).value  # i.e. cell H7
        u_denominator = worksheet.cell(7, 9).value  # i.e. cell I7
    elif data_type == 'LIST':
        row_classifications = pd.read_excel(file, sheet_name='Cover', usecols='F:G',
                                            skiprows=[i for i in range(10)],
                                            index_col="Aspects_classifications").dropna()
        # Rename values column so it has the same name as the TABLE type. The two templates should be harmonized.
        row_classifications = row_classifications.rename({'Aspects_Attribute_No': 'Row_Aspects_Attribute_No'})
        col_classifications = 'LIST'
        data_info = pd.read_excel(file, sheet_name='Cover', usecols='H:I',
                                  skiprows=[i for i in range(10)],
                                  index_col="DATA").dropna()
        u_nominator = 'LIST'
        u_denominator = 'LIST'
    else:
        raise AssertionError("Unknown data type or malformed Excel file. Cell Cover!G10 should be 'LIST' or 'TABLE',"
                             " but is '%s'" % data_type)
    workbook.close()
    return {'data_type': data_type,
            'dataset_info': dataset_info,
            'data_sources': data_sources,
            'row_classifications': row_classifications,
            'col_classifications': col_classifications,
            'data_info': data_info,
            'u_nominator': u_nominator,
            'u_denominator': u_denominator}


def read_candidate_data_list(file, path=IEDC_paths.candidates):
    """
    Will read a candidate file and return its data.

    :param file: Filename of the file to process
    :param path: Path of the file
    :return: Dictionary of dataframes for metadata, classifications, and data
    """
    # make it a proper path
    file = os.path.join(path, file)
    data = pd.read_excel(file, sheet_name='Data')
    return data


def read_candidate_data_table(file, aspects_table, path=IEDC_paths.candidates):
    """
    Will read a candidate file and return its data.

    :param file: Filename of the file to process
    :param path: Path of the file
    :return: Dictionary of dataframes for metadata, classifications, and data
    """
    row_indices = aspects_table[aspects_table['position'].str.startswith('row')].sort_values('position')['name']
    col_indices = aspects_table[aspects_table['position'].str.startswith('col')].sort_values('position')['name']
    # make it a proper path
    file = os.path.join(path, file)
    data = pd.read_excel(file, sheet_name='Data', header=[i for i in range(len(col_indices))],
                         index_col=[i for i in range(len(row_indices))])
    # Check that same column names are not interpreted once as int and once as str
    # https://stackoverflow.com/a/54393368/2075003
    if hasattr(data.columns, 'levels'):
        # Weird workaroud required https://stackoverflow.com/questions/54393343/harmonise-pandas-multiindex-to-string-when-reading-excel-files/54393368#comment98729305_54393368
        old_df_codes = data.columns.codes
        data.columns = pd.MultiIndex.from_product([clevel.astype(str).unique() for clevel in data.columns.levels])
        data.columns.set_codes(old_df_codes, inplace=True)
    else:
        data.columns = data.columns.astype(str).unique()
    data.columns.names = col_indices
    data.index.names = row_indices
    return data


def read_units_table(file, row_indices, col_indices, path=IEDC_paths.candidates):
    # file = os.path.join(path, 'TABLE', file)
    file = os.path.join(path, file)
    units = {}
    for u in ['Unit_nominator', 'Unit_denominator']:
        units[u] = pd.read_excel(file, sheet_name=u, header=[i for i in range(len(col_indices))],
                                 index_col=[i for i in range(len(row_indices))])
        units[u].columns.names = col_indices
        units[u].index.names = row_indices
    return units


def read_stats_array_table(file, row_indices, col_indices, path=IEDC_paths.candidates):
    # file = os.path.join(path, 'TABLE', file)
    file = os.path.join(path, file)
    sa_df = pd.read_excel(file, sheet_name='stats_array_string', header=[i for i in range(len(col_indices))],
                          index_col=[i for i in range(len(row_indices))])
    sa_df.columns.names = col_indices
    sa_df.index.names = row_indices
    return sa_df


def read_comment_table(file, row_indices, col_indices, path=IEDC_paths.candidates):
    # file = os.path.join(path, 'TABLE', file)
    file = os.path.join(path, file)
    sa_df = pd.read_excel(file, sheet_name='Comment', header=[i for i in range(len(col_indices))],
                          index_col=[i for i in range(len(row_indices))])
    sa_df.columns.names = col_indices
    sa_df.index.names = row_indices
    return sa_df


def read_candidate_files(path=IEDC_paths.candidates):
    """
    Runs read_candidate_file() for all files in a directory
    :param path: directory of files to run the function for
    :return: TODO: List of dataframes?
    """
    for file in get_candidate_filenames(path):
        read_candidate_data_list(file, path)
    # TODO
    return None


def ds_in_db(file_meta, crash=True):
    """
    Checks if a dataset is already in the database
    """
    db_datasets = dbio.get_sql_table_as_df('datasets')
    indb = db_datasets[['dataset_name', 'dataset_version']].values.tolist()
    ds_name = file_meta['dataset_info'].loc['dataset_name'].values[0]
    ds_ver = file_meta['dataset_info'].loc['dataset_version'].values[0]
    if ds_ver is np.nan:
        ds_ver = None
    else:
        ds_ver = str(ds_ver)
    candidate = [ds_name, ds_ver]
    if candidate in indb:
        if crash:
            raise AssertionError("Dataset already in DB %s" % candidate)
        print("Dataset already in DB %s" % candidate)
        return True
    else:
        return False

